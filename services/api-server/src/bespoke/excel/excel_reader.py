import openpyxl
import logging

from typing import Any, List, Tuple, Dict
from mypy_extensions import TypedDict

from bespoke.excel.excel_types import CellValue

SheetDict = TypedDict('SheetDict', {
	'sheet_name': str,
	'rows': List[List[CellValue]]
})

WorkbookDict = TypedDict('WorkbookDict', {
	'sheets': List[SheetDict],
	'sheet_names': List[str]
})

def _to_sheet_obj(workbook: Any, sheet_name: str, trim_empty_columns: bool) -> Tuple[SheetDict, str]:
	try:
		sheet = workbook[sheet_name]
	except Exception as e:
		return None, u'Reading from Excel error: {0}'.format(e)

	sheet_obj = SheetDict(rows=[], sheet_name=sheet_name)
	for row in sheet.iter_rows():
		return_row: List[CellValue] = []
		for cell in row:
			if not cell.value:
				return_row.append('')
				continue

			if type(cell.value) == str:
				return_row.append('{}'.format(cell.value).strip(''))
			else:
				return_row.append(cell.value)

		sheet_obj['rows'].append(return_row)

	return sheet_obj, None

class ExcelWorkbook(object):
	
	def __init__(self, private: bool) -> None:
		self.d  = WorkbookDict(
			sheets=[],
			sheet_names=[]
		)
		self._index_to_name: Dict[int, str] = {}

	def get_sheet_names(self) -> List[str]:
		return self.d['sheet_names']

	def add_sheet(self, sheet_obj: SheetDict) -> None:
		self.d['sheets'].append(sheet_obj)

	def get_sheet_by_name(self, name: str) -> Tuple[SheetDict, str]:
		for sheet in self.d['sheets']:
			if sheet['sheet_name'] == name:
				return sheet, None
				
		return None, u'No sheet found named {}'.format(name)

	def get_sheet_by_index(self, index: int) -> Tuple[SheetDict, str]:
		num_sheets = len(self.d['sheets'])
		if index < 0 or index >= num_sheets:
			return None, u'Invalid index to get_sheet_by_index. Must be between 0 and {}'.format(num_sheets)
			
		return self.d['sheets'][index], None 

	@staticmethod
	def load_xlsx(filename: str, trim_empty_columns: bool = False) -> Tuple['ExcelWorkbook', str]:
		excel = ExcelWorkbook(private=True)
		try:
			workbook = openpyxl.load_workbook(filename=filename, data_only=True)
			sheet_names = workbook.sheetnames
			excel.d['sheet_names'] = sheet_names
			num_sheets = len(sheet_names)
			
			for i in range(num_sheets):
				sheet_name = sheet_names[i]
				excel._index_to_name[i] = sheet_name
				sheet_obj, err = _to_sheet_obj(workbook, sheet_name, trim_empty_columns)
				if err:
					return None, err
				excel.add_sheet(sheet_obj)
				
			return excel, None
		except Exception as ex:
			logging.error(ex)
			return None, u'Error processing the excel file {}'.format(ex)

if __name__ == '__main__':
	path = '/Users/davidlluncor/Desktop/Buddies_Accounting_Master_20201102.xlsx'

	workbook, err = ExcelWorkbook.load_xlsx(path)
	if err:
		raise Exception(err)

	sheet_names = workbook.get_sheet_names()
	print(sheet_names)

	summary, err = workbook.get_sheet_by_index(0)
	if err:
		raise Exception(err)
	print(summary)

	for sheet_name in sheet_names:
		print('==========')
		sheet, err = workbook.get_sheet_by_name(sheet_name)
		if err:
			raise Exception(err)
		print(sheet)
